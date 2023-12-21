import websocket
import json
import threading
import time
import random
import math
import hashlib
import csv
import pandas as pd 
import os
from cryptography.hazmat.primitives import hashes
from cryptography.hazmat.primitives.asymmetric import rsa, padding
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

VERIFICATION_THRESHOLD = 0.7  # 70% of True votes required for successful verification

# Constants
NUMBER_OF_NODES = 10
FAULTY_PROPORTION = 1/3
NUMBER_OF_FAULTY_NODES = math.floor(NUMBER_OF_NODES * FAULTY_PROPORTION)
CHUNK_SIZE = 500  # Adjusted chunk size
BUFFER_CHECK_FREQUENCY = 20  # Seconds
DATA_BUFFER = ""  # Initialize data buffer
exit_flag = False  # Flag for graceful exit

NUMBER_OF_SEGMENTS = 5  # Define the number of segments per chunk
NUMBER_OF_CHUNKS = 3    # Define the number of chunks to process

import os

def save_to_excel(chunk_number, timestamp, chunk_data, chunk_size, digital_signature, outcome, true_vote_percentage):
    # Create a DataFrame for the chunk data
    df = pd.DataFrame({
        "Chunk #": [chunk_number],
        "Timestamp": [timestamp],
        "Chunk Data": [chunk_data],
        "Chunk Size": [chunk_size],
        "Digital Signature": [digital_signature],
        "Verification Outcome": [outcome],
        "True Vote Percentage": [true_vote_percentage]
    })
    
    # Define the Excel writer
    excel_file = 'data.xlsx'
    sheet_name = 'Sheet1'
    
    # Check if the Excel file already exists
    if not os.path.isfile(excel_file):
        with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
    else:
        # If the file exists, then open it and append data without adding header
        with pd.ExcelWriter(excel_file, mode='a', engine='openpyxl', if_sheet_exists='overlay') as writer:
            # Get the last row in the existing Excel sheet
            # If there is no sheet, then index is 0 by default
            startrow = writer.sheets[sheet_name].max_row if sheet_name in writer.sheets else 0
            
            # Write data starting from the last row
            df.to_excel(writer, sheet_name=sheet_name, startrow=startrow, index=False, header=False)

# Function to print colored text in the terminal
def print_colored(text, color):
    colors = {
        'green': '\033[92m',
        'red': '\033[91m',
        'end': '\033[0m',
    }
    print(f"{colors[color]}{text}{colors['end']}")

# Generate RSA keys (private and public)
private_key = rsa.generate_private_key(public_exponent=65537, key_size=2048)
public_key = private_key.public_key()

# CSV File Initialization
csv_filename = 'data.csv'
with open(csv_filename, 'w', newline='') as file:
    writer = csv.writer(file)
    writer.writerow(["Chunk #", "Timestamp", "Chunk Data", "Chunk Size", "Digital Signature", "Verification Outcome", "True Vote Percentage"])

# Node Class
class Node:
    def __init__(self, node_id, is_faulty):
        self.node_id = node_id
        self.is_faulty = is_faulty

    def vote(self, segment, segment_hash, public_key, signature, timestamp):
        if self.is_faulty:
            return random.choice([True, False])

        calculated_hash = hashlib.sha256(segment.encode()).hexdigest()
        if calculated_hash != segment_hash:
            return False

        try:
            public_key.verify(
                signature,
                calculated_hash.encode(),
                padding.PSS(
                    mgf=padding.MGF1(hashes.SHA256()),
                    salt_length=padding.PSS.MAX_LENGTH
                ),
                hashes.SHA256()
            )
        except Exception:
            return False

        if datetime.now().timestamp() - timestamp > 300:
            return False

        return True

# DHT Class
class DHT:
    def __init__(self, public_key):
        self.nodes = [Node(i, i < NUMBER_OF_FAULTY_NODES) for i in range(NUMBER_OF_NODES)]
        self.public_key = public_key

    def process_segment(self, segment, segment_hash, signature, timestamp):
        return {node.node_id: node.vote(segment, segment_hash, self.public_key, signature, timestamp) for node in self.nodes}

# Function to segment the data
# Function to segment the data
def segment_data(data, private_key):
    head = data[:CHUNK_SIZE // (NUMBER_OF_SEGMENTS + 2)]
    tail = data[-CHUNK_SIZE // (NUMBER_OF_SEGMENTS + 2):]
    segments = [head] + [data[i:i + CHUNK_SIZE // NUMBER_OF_SEGMENTS] for i in range(CHUNK_SIZE // (NUMBER_OF_SEGMENTS + 2), len(data) - CHUNK_SIZE // (NUMBER_OF_SEGMENTS + 2), CHUNK_SIZE // NUMBER_OF_SEGMENTS)] + [tail]

    segment_info = []
    for segment in segments:
        segment_hash = hashlib.sha256(segment.encode()).hexdigest()
        signature = private_key.sign(
            segment_hash.encode(),
            padding.PSS(
                mgf=padding.MGF1(hashes.SHA256()),
                salt_length=padding.PSS.MAX_LENGTH
            ),
            hashes.SHA256()
        )
        timestamp = datetime.now().timestamp()
        segment_info.append((segment, segment_hash, signature, timestamp))

    return segment_info

# Function to save chunk data to CSV
def save_to_csv(chunk_number, timestamp, chunk_data, chunk_size, digital_signature, outcome, true_vote_percentage):
    with open(csv_filename, 'a', newline='') as file:
        writer = csv.writer(file)
        writer.writerow([chunk_number, timestamp, chunk_data, chunk_size, digital_signature, outcome, true_vote_percentage])

# Function to process and clear the buffer
def process_buffer(dht, chunk_number):
    global DATA_BUFFER, VERIFICATION_THRESHOLD
    if len(DATA_BUFFER) >= CHUNK_SIZE:
        chunk_data = DATA_BUFFER[:CHUNK_SIZE]
        DATA_BUFFER = DATA_BUFFER[CHUNK_SIZE:]
        print(f"Processing Chunk {chunk_number}...")
        segments_info = segment_data(chunk_data, private_key)
        node_votes = {node_id: [] for node_id in range(NUMBER_OF_NODES)}
        chunk_consensus = True
        total_true_votes = 0

        for segment_index, (segment, segment_hash, signature, timestamp) in enumerate(segments_info):
            segment_votes = dht.process_segment(segment, segment_hash, signature, timestamp)
            for node_id, vote in segment_votes.items():
                node_votes[node_id].append(vote)
                total_true_votes += int(vote)

        print(f"Chunk {chunk_number}")
        for node_id, votes in node_votes.items():
            print(f"Node {node_id} Votes: {votes}")

        # Calculate the consensus and true vote percentage
        total_votes = NUMBER_OF_NODES * (NUMBER_OF_SEGMENTS + 2)
        true_vote_percentage = (total_true_votes / total_votes) * 100
        chunk_consensus = true_vote_percentage >= (VERIFICATION_THRESHOLD * 100)

        # Print results with colored text
        consensus_text = 'Verified Successfully' if chunk_consensus else 'Verified Unsuccessfully'
        print_colored(f"Chunk {chunk_number} {consensus_text} ({true_vote_percentage:.2f}% true)", 'green' if chunk_consensus else 'red')

        # Compute digital signature for the entire chunk
        chunk_signature = private_key.sign(
            chunk_data.encode(),
            padding.PSS(
                mgf=padding.MGF1(hashes.SHA256()),
                salt_length=padding.PSS.MAX_LENGTH
            ),
            hashes.SHA256()
        )
        chunk_signature_readable = chunk_signature.hex()

        # Save chunk data to Excel
        save_to_excel(chunk_number, datetime.now().timestamp(), chunk_data, len(chunk_data), chunk_signature_readable, consensus_text, true_vote_percentage)

# Function to check buffer size periodically
def check_buffer_size(dht):
    global exit_flag
    chunk_number = 1
    while not exit_flag and chunk_number <= NUMBER_OF_CHUNKS:
        if len(DATA_BUFFER) >= CHUNK_SIZE:
            process_buffer(dht, chunk_number)
            if chunk_number == NUMBER_OF_CHUNKS:
                exit_flag = True  # Set exit flag to ensure main program exits
                break
            chunk_number += 1
        else:
            print(f"Waiting for enough data for Chunk {chunk_number}...")
        time.sleep(BUFFER_CHECK_FREQUENCY)
    print("Exiting buffer check thread.")

# Binance WebSocket message processing function
def process_binance_message(ws_message):
    global DATA_BUFFER
    message_data = json.loads(ws_message)
    if 'p' in message_data:
        DATA_BUFFER += message_data['p']
    else:
        print("No price field in the received message.")

# Binance WebSocket event handlers
def on_message(ws, message):
    process_binance_message(message)

def on_error(ws, error):
    print("Error:", error)

def on_close(ws, close_status_code, close_msg):
    print("WebSocket closed")

def on_open(ws):
    print("WebSocket connection opened to Binance")

# Set up Binance WebSocket connection and run client
binance_ws_url = "wss://stream.binance.com:9443/ws/btcusdt@trade"
ws = websocket.WebSocketApp(binance_ws_url, on_message=on_message, on_error=on_error, on_close=on_close)

dht = DHT(public_key)
ws_thread = threading.Thread(target=ws.run_forever)
ws_thread.start()

buffer_thread = threading.Thread(target=lambda: check_buffer_size(dht))
buffer_thread.start()

# Main loop for keeping the script running
try:
    while True:
        if exit_flag:
            break
        time.sleep(0.1)
except KeyboardInterrupt:
    print("Exiting program...")
    exit_flag = True

# Close WebSocket and wait for threads to finish
ws.close()
buffer_thread.join()
ws_thread.join()

print("Program exited gracefully.")
